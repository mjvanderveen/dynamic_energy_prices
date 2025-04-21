import requests
import json
from datetime import datetime, timedelta
import urllib.parse
import os
import csv
import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import random

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the full path to config.json
config_path = os.path.join(script_dir, "config.json")

# Load configuration from config.json
with open(config_path, "r") as config_file:
    config = json.load(config_file)

# Configuration variables
DYNAMIC_PRICES_API_URL = config["DATA"]["DYNAMIC_PRICES_API_URL"]
DYNAMIC_PRICES_API_KEY = config["DATA"]["DYNAMIC_PRICES_API_KEY"]
START_DATE = config["PARAMETERS"]["START_DATE"]
END_DATE = config["PARAMETERS"]["END_DATE"]
CONSUMPTION_SENSORS = config["CONSUMPTION_SENSORS"]
PRODUCTION_SENSORS = config["PRODUCTION_SENSORS"]
VICTORIAMETRICS_URL = config["DATA"]["VICTORIAMETRICS_URL"]

# Load taxes
TAXES = config["TAXES"]
ENERGY_TAX = TAXES["ENERGY_TAX"]  # Energy tax per kWh (in euro)
STORAGE_COSTS = TAXES["STORAGE_COSTS"]  # Storage costs per kWh (in euro)
STORAGE_COSTS_PRODUCTION = TAXES["STORAGE_COSTS_PRODUCTION"]  # Storage costs for production (in euro)
VAT = TAXES["VAT"]  # VAT percentage
FIXED_SUPPLY_COSTS = TAXES["FIXED_SUPPLY_COSTS"]  # Fixed supply costs per month (in euro)
TRANSPORT_COSTS = TAXES["TRANSPORT_COSTS"]  # Transport costs per month (in euro)
ENERGY_TAX_COMPENSATION = TAXES["ENERGY_TAX_COMPENSATION"]  # Energy tax compensation per month (in euro)

# Load debug setting
DEBUG = config["PARAMETERS"].get("DEBUG", False)

# Load production stop setting
STOP_PRODUCTION_NEGATIVE_PRICES = config["PARAMETERS"].get("STOP_PRODUCTION_NEGATIVE_PRICES", False)
def debug_print(message):
    """Print debug messages if DEBUG is enabled."""
    if DEBUG:
        print(message)

def fetch_sensor_data_from_json(file_path, start_date, end_date, sensor_ids, output_file=None):
    """
    Fetch and parse sensor data from a JSON file (export.json), filtering by sensor IDs and date range.

    Args:
        file_path (str): Path to the JSON file.
        start_date (str): The start date for filtering data (format: YYYY-MM-DD).
        end_date (str): The end date for filtering data (format: YYYY-MM-DD).
        sensor_ids (list): List of sensor IDs to filter the data.
        output_file (str): Path to save the raw filtered data as a CSV file (optional).

    Returns:
        dict: A dictionary with timestamps as keys and hourly sensor increments as values.
    """
    try:
        # Load the JSON file
        with open(file_path, "r") as file:
            data = json.load(file)

        # Convert start_date and end_date to datetime objects
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

        # Initialize the filtered data list
        filtered_data = []

        # Process each sensor ID separately
        for sensor_id in sensor_ids:
            # Filter the data for the current sensor ID
            sensor_data = [
                record for record in data
                if record["statistic_id"] == sensor_id
                and start_datetime <= datetime.strptime(record["d"], "%Y-%m-%d %H:%M:%S") < end_datetime
            ]

            # Add the filtered data for the current sensor to the overall filtered data
            filtered_data.extend(sensor_data)

        # Debug: Print the number of records fetched
        debug_print(f"Filtered {len(filtered_data)} records from {file_path} for sensors: {sensor_ids} within date range {start_date} to {end_date}")

        # Save the raw filtered data to a CSV file if output_file is provided
        if output_file:
            try:
                # Ensure the data folder exists
                data_folder = os.path.join(script_dir, "data")
                os.makedirs(data_folder, exist_ok=True)

                # Construct the full path for the output file
                output_file_path = os.path.join(data_folder, output_file)

                # Write the filtered data to the CSV file
                with open(output_file_path, mode="w", newline="") as csv_file:
                    writer = csv.writer(csv_file)
                    # Write the header
                    writer.writerow(["statistic_id", "timestamp", "increment"])
                    # Write the data
                    for record in filtered_data:
                        writer.writerow([record["statistic_id"], record["d"], record["increment"]])
                debug_print(f"Raw filtered data written to {output_file_path}")
            except IOError as e:
                debug_print(f"Failed to write raw filtered data to {output_file}: {e}")

        # Process the cumulative data
        return process_cumulative_data(filtered_data, start_date, end_date, "d", "increment", "%Y-%m-%d %H:%M:%S")
    except (IOError, json.JSONDecodeError) as e:
        print(f"Error reading or parsing JSON file {file_path}: {e}")
        return {}
    
def fetch_sensor_data_victoriametrics(sensor_ids, start_date, end_date, output_file):
    """
    Fetch historical sensor data from VictoriaMetrics using the delta function to calculate increments.

    Args:
        sensor_ids (list): List of sensor IDs to query.
        start_date (str): The start date for the query (format: YYYY-MM-DDT00:00:00Z).
        end_date (str): The end date for the query (format: YYYY-MM-DDT23:59:59Z).
        output_file (str): Path to save the combined raw data.

    Returns:
        dict: A dictionary with timestamps as keys and hourly sensor values as values.
    """
    # Convert start_date and end_date to Unix timestamps (VictoriaMetrics requires this format)
    start_datetime = datetime.strptime(start_date, "%Y-%m-%dT%H:%M:%SZ")
    end_datetime = datetime.strptime(end_date, "%Y-%m-%dT%H:%M:%SZ")
    start_timestamp = int(start_datetime.timestamp())
    end_timestamp = int(end_datetime.timestamp())

    combined_data = []  # List to store combined raw data for all sensors
    hourly_totals = {}  # Dictionary to store hourly increments

    for sensor_id in sensor_ids:
        # Use the delta function in VictoriaMetrics to calculate increments
        query = f'delta({sensor_id}_value[1h])'
        params = {
            "query": query,
            "start": start_timestamp,
            "end": end_timestamp,
            "step": "3600s"
        }

        response = requests.get(VICTORIAMETRICS_URL, params=params)

        if response.status_code == 200:
            # Parse the response
            raw_data = response.json().get("data", {}).get("result", [])
            for result in raw_data:
                for ts, val in result.get("values", []):
                    # Convert the timestamp to UTC datetime
                    utc_timestamp = datetime.utcfromtimestamp(int(ts))

                    # Calculate summer and winter time transitions
                    year = utc_timestamp.year

                    # Correct calculation for the last Sunday of March (DST Start)
                    dst_start = (
                        datetime(year, 4, 1) - timedelta(days=(datetime(year, 4, 1).weekday() + 1))
                    ).replace(hour=2)  # DST starts at 02:00 AM CET

                    # Correct calculation for the last Sunday of October (DST End)
                    dst_end = (
                        datetime(year, 11, 1) - timedelta(days=(datetime(year, 11, 1).weekday() + 1))
                    ).replace(hour=3)  # DST ends at 03:00 AM CEST

                    # Shift timestamps if they fall within DST
                    if dst_start <= utc_timestamp < dst_end:
                        adjusted_timestamp = utc_timestamp + timedelta(hours=1)
                    else:
                        adjusted_timestamp = utc_timestamp

                    # Format the adjusted timestamp as YYYY-MM-DDTHH
                    formatted_timestamp = adjusted_timestamp.strftime("%Y-%m-%dT%H")

                    # Store the increment value
                    increment = float(val)
                    hourly_totals[formatted_timestamp] = increment

                    # Add the raw data to combined_data
                    combined_data.append({
                        "statistic_id": sensor_id,
                        "d": formatted_timestamp,
                        "value": increment
                    })
        else:
            debug_print(f"Failed to fetch data for {sensor_id}: {response.status_code}, {response.text}")

    # Save combined raw data to a JSON file
    try:
        # Ensure the data folder exists
        data_folder = os.path.join(script_dir, "data")
        os.makedirs(data_folder, exist_ok=True)

        # Construct the full path for the output file
        output_file_path = os.path.join(data_folder, output_file)

        # Write the combined data to the file
        with open(output_file_path, "w") as file:
            json.dump(combined_data, file, indent=4)
        debug_print(f"Combined raw data written to {output_file_path}")
    except IOError as e:
        debug_print(f"Failed to write combined raw data to {output_file_path}: {e}")

    return hourly_totals

def process_cumulative_data(data, start_date, end_date, timestamp_key, value_key, timestamp_format):
    """
    Process cumulative data to aggregate hourly values.

    Args:
        data (list): List of records containing hourly increments.
        start_date (str): The start date for filtering data (format: YYYY-MM-DD).
        end_date (str): The end date for filtering data (format: YYYY-MM-DD).
        timestamp_key (str): The key in the record that contains the timestamp.
        value_key (str): The key in the record that contains the increment value.
        timestamp_format (str): The format of the timestamp in the data.

    Returns:
        dict: A dictionary with timestamps as keys and aggregated hourly values as values.
    """
    # Convert start_date and end_date to datetime objects
    start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
    end_datetime = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

    # Initialize the result dictionary
    hourly_totals = {}

    # Process each record in the data
    for record in data:
        try:
            # Parse the timestamp and value
            timestamp = datetime.strptime(record[timestamp_key], timestamp_format)
            value = float(record[value_key])

            # Filter records within the specified date range
            if start_datetime <= timestamp < end_datetime:
                # Format the timestamp as YYYY-MM-DDTHH
                formatted_timestamp = timestamp.strftime("%Y-%m-%dT%H")

                # Aggregate the value
                if formatted_timestamp not in hourly_totals:
                    hourly_totals[formatted_timestamp] = 0
                hourly_totals[formatted_timestamp] += value
        except (KeyError, ValueError) as e:
            debug_print(f"Skipping invalid record: {record}, Error: {e}")

    return hourly_totals

def write_hourly_comparison_to_csv(victoriametrics_data, export_json_data, output_file):
    """
    Write a CSV file comparing hourly kWh data from VictoriaMetrics and export.json.

    Args:
        victoriametrics_data (dict): Hourly kWh data from VictoriaMetrics.
        export_json_data (dict): Hourly kWh data from export.json.
        output_file (str): Path to save the comparison CSV file.
    """
    # Ensure the results folder exists
    results_folder = "results"
    os.makedirs(results_folder, exist_ok=True)

    # Construct the full path for the output file
    output_file_path = os.path.join(results_folder, output_file)

    # Get all unique timestamps from both data sources
    all_timestamps = set(victoriametrics_data.keys()).union(set(export_json_data.keys()))

    # Write the comparison data to the CSV file
    with open(output_file_path, mode="w", newline="") as csv_file:
        writer = csv.writer(csv_file)

        # Write the header
        writer.writerow(["Timestamp", "VictoriaMetrics (kWh)", "Export.json (kWh)"])

        # Write the data for each timestamp
        for timestamp in sorted(all_timestamps):
            victoriametrics_value = victoriametrics_data.get(timestamp, 0)
            export_json_value = export_json_data.get(timestamp, 0)
            writer.writerow([timestamp, f"{victoriametrics_value:.3f}", f"{export_json_value:.3f}"])

    print(f"Hourly comparison written to {output_file_path}")

def fetch_dynamic_prices(start_date, end_date):
    """Fetch dynamic energy prices for the given date range, handling multiple years and caching."""
    # Parse the start and end years
    start_year = datetime.strptime(start_date, "%Y-%m-%d").year
    end_year = datetime.strptime(end_date, "%Y-%m-%d").year
    current_date = datetime.now().date()

    # Initialize an empty list to store price data
    combined_price_data = []

    # Loop through each year in the range
    for year in range(start_year, end_year + 1):
        cache_file = f"./data/dynamic_energy_prices_{year}.json"  # Cache file for the year

        # Determine if the year is in the past or the current year
        if year < current_date.year:
            # For past years, always use the cached data if available
            if os.path.exists(cache_file):
                try:
                    with open(cache_file, "r") as file:
                        cached_data = json.load(file)
                        debug_print(f"Using cached dynamic prices for year {year}")
                        combined_price_data.extend(normalize_price_data(cached_data))
                        continue
                except (json.JSONDecodeError, IOError):
                    debug_print(f"Failed to read cache file for year {year}, fetching from API...")

        elif year == current_date.year:
            # For the current year, check if the cache is up-to-date (download once per day)
            if os.path.exists(cache_file):
                last_modified = datetime.fromtimestamp(os.path.getmtime(cache_file)).date()
                if last_modified == current_date:
                    try:
                        with open(cache_file, "r") as file:
                            cached_data = json.load(file)
                            debug_print(f"Using cached dynamic prices for year {year} (up-to-date)")
                            combined_price_data.extend(normalize_price_data(cached_data))
                            continue
                    except (json.JSONDecodeError, IOError):
                        debug_print(f"Failed to read cache file for year {year}, fetching from API...")

        # Fetch data from the API if cache is not available or outdated
        url = f"{DYNAMIC_PRICES_API_URL}?period=jaar&year={year}&type=json&key={DYNAMIC_PRICES_API_KEY}"
        debug_print(f"Fetching dynamic prices from: {url}")
        response = requests.get(url)
        if response.status_code == 200:
            price_data = json.loads(response.text)

            # Save the fetched data to the cache file
            try:
                with open(cache_file, "w") as file:
                    json.dump(price_data, file)
                    print(f"Cached dynamic prices for year {year}")
            except IOError:
                print(f"Failed to write cache file for year {year}")

            combined_price_data.extend(normalize_price_data(price_data))
        else:
            debug_print(f"Failed to fetch dynamic prices from API for year {year}: {response.status_code}")

    return combined_price_data

def normalize_price_data(price_data):
    """Normalize the timestamps in price_data to the format YYYY-MM-DDTHH."""
    normalized_data = []
    for entry in price_data:
        try:
            # Parse the timestamp and reformat it to YYYY-MM-DDTHH
            original_timestamp = entry["datum"]
            # Handle both formats: with 'T' or with a space
            if "T" in original_timestamp:
                normalized_timestamp = datetime.strptime(original_timestamp, "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%dT%H")
            else:
                normalized_timestamp = datetime.strptime(original_timestamp, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%dT%H")
            entry["datum"] = normalized_timestamp
            normalized_data.append(entry)
        except (KeyError, ValueError) as e:
            debug_print(f"Invalid price entry: {entry}, Error: {e}")
    return normalized_data

def simulate_battery_with_forecasting(
    hourly_consumption,
    hourly_production,
    battery_state,
    config,
    future_prices,
    historical_consumption,
    future_production_forecast,
    timestamp
):
    """
    Simulate the behavior of a battery using future price data and forecasting.

    Args:
        hourly_consumption (float): The energy consumption for the hour (kWh).
        hourly_production (float): The energy production for the hour (kWh).
        battery_state (dict): The current state of the battery.
        config (dict): The battery configuration.
        future_prices (list): Future price data for the next day.
        historical_consumption (dict): Historical consumption data for forecasting.
        future_production_forecast (dict): Forecasted production data for the next day.
        timestamp (str): The timestamp for the current hour (format: YYYY-MM-DDTHH).

    Returns:
        tuple: Adjusted consumption, adjusted production, updated battery state, energy loss.
    """
    # Extract battery parameters from the config
    battery_size = config["BATTERY_SIMULATION"]["BATTERY_SIZE_KWH"]
    max_charging_rate = config["BATTERY_SIMULATION"]["MAX_CHARGING_RATE_KWH"]
    max_discharging_rate = config["BATTERY_SIMULATION"]["MAX_DISCHARGING_RATE_KWH"]
    round_trip_efficiency = config["BATTERY_SIMULATION"].get("ROUND_TRIP_EFFICIENCY", 0.8)
    discharge_minimum = (config["BATTERY_SIMULATION"]["DISCHARGE_MINIMUM_PERCENTAGE"] / 100) * battery_size
    charge_maximum = (config["BATTERY_SIMULATION"]["CHARGE_MAXIMUM_PERCENTAGE"] / 100) * battery_size
    error_rate = config["BATTERY_SIMULATION"]["PRODUCTION_FORECAST_ERROR_RATE"] / 100

    # Get the current battery level
    battery_level = battery_state["level"]

    # Initialize charge/discharge tracking
    charge_amount = 0
    discharge_amount = 0
    energy_loss = 0  # Track energy loss due to round-trip efficiency
    simulated_consumption = hourly_consumption
    simulated_production = hourly_production

    # Forecast high and low consumption hours using historical data
    forecasted_high_consumption_hours = [
        hour for hour, value in historical_consumption.items() if value > (sum(historical_consumption.values()) / len(historical_consumption))
    ]

    # Apply a random error rate to future production forecasts
    forecasted_production = {
        hour: value * (1 + random.uniform(-error_rate, error_rate))
        for hour, value in future_production_forecast.items()
    }

    # Get the price for the current hour
    current_price = future_prices.get(timestamp, {"price_consumption": 0, "price_production": 0})
    price_consumption = current_price["price_consumption"]
    price_production = current_price["price_production"]

    # Case 1: Charge the battery if the price is low
    if price_production < config["BATTERY_SIMULATION"]["DYNAMIC_PRICE_THRESHOLD_LOW"]:
        charge_amount = min(max_charging_rate, charge_maximum - battery_level)
        if charge_amount > 0:
            battery_level += charge_amount * round_trip_efficiency
            energy_loss += charge_amount * (1 - round_trip_efficiency)
            battery_state["total_charged"] += charge_amount
            simulated_production -= charge_amount  # Reduce production used for charging

    # Case 2: Discharge the battery for consumption if the consumption price is high
    elif timestamp in forecasted_high_consumption_hours and price_consumption > config["BATTERY_SIMULATION"]["DYNAMIC_PRICE_THRESHOLD_HIGH"]:
        discharge_amount = min(hourly_consumption, max_discharging_rate, max(0, battery_level - discharge_minimum))
        if discharge_amount > 0:
            battery_level -= discharge_amount
            simulated_consumption -= discharge_amount
            battery_state["total_discharged"] += discharge_amount

    # Case 3: Discharge the battery for production if the production price is high
    elif price_production > config["BATTERY_SIMULATION"]["DYNAMIC_PRICE_THRESHOLD_HIGH"]:
        discharge_amount = min(max_discharging_rate, max(0, battery_level - discharge_minimum))
        if discharge_amount > 0:
            battery_level -= discharge_amount
            simulated_production += discharge_amount
            battery_state["total_discharged"] += discharge_amount

    # Update the battery state
    battery_state["level"] = battery_level

    # Calculate charge cycles
    usable_capacity = charge_maximum - discharge_minimum
    battery_state["charge_cycles"] = int(battery_state["total_discharged"] // usable_capacity)

    return simulated_consumption, simulated_production, battery_state, energy_loss

def simulate_battery(hourly_consumption, hourly_production, battery_state, config, total_price_incl_vat_production, total_price_incl_vat_consumption, timestamp, strategy):
    """
    Simulate the behavior of a battery for a single hour based on the chosen strategy.

    Args:
        hourly_consumption (float): The energy consumption for the hour (kWh).
        hourly_production (float): The energy production for the hour (kWh).
        battery_state (dict): The current state of the battery.
        config (dict): The battery configuration.
        total_price_incl_vat_production (float): The energy price for production (including taxes).
        timestamp (str): The timestamp for the current hour (format: YYYY-MM-DDTHH).
        strategy (str): The battery charge strategy ("self-sufficiency" or "dynamic_cost_optimization").

    Returns:
        tuple: Adjusted consumption, adjusted production, updated battery state, energy loss.
    """
    # Extract battery parameters from the config
    battery_size = config["BATTERY_SIMULATION"]["BATTERY_SIZE_KWH"]
    max_charging_rate = config["BATTERY_SIMULATION"]["MAX_CHARGING_RATE_KWH"]
    max_discharging_rate = config["BATTERY_SIMULATION"]["MAX_DISCHARGING_RATE_KWH"]
    round_trip_efficiency = config["BATTERY_SIMULATION"].get("ROUND_TRIP_EFFICIENCY", 0.8)
    discharge_minimum = (config["BATTERY_SIMULATION"]["DISCHARGE_MINIMUM_PERCENTAGE"] / 100) * battery_size
    charge_maximum = (config["BATTERY_SIMULATION"]["CHARGE_MAXIMUM_PERCENTAGE"] / 100) * battery_size

    # Get the current battery level
    battery_level = battery_state["level"]

    # Initialize charge/discharge tracking
    charge_amount = 0
    discharge_amount = 0
    energy_loss = 0  # Track energy loss due to round-trip efficiency
    simulated_consumption = hourly_consumption
    simulated_production = hourly_production

    if strategy == "self-sufficiency":
        # Self-Sufficiency Strategy
        # Adjust production by charging the battery
        if hourly_production > 0:
            charge_amount = min(hourly_production, max_charging_rate, charge_maximum - battery_level)
            if charge_amount > 0:
                battery_level += charge_amount * round_trip_efficiency
                simulated_production -= charge_amount
                energy_loss += charge_amount * (1 - round_trip_efficiency)
                battery_state["total_charged"] += charge_amount

        # Adjust consumption by discharging the battery
        if hourly_consumption > 0:
            discharge_amount = min(hourly_consumption, max_discharging_rate, battery_level - discharge_minimum)
            if discharge_amount > 0:
                battery_level -= discharge_amount
                simulated_consumption -= discharge_amount
                battery_state["total_discharged"] += discharge_amount

        # Handle excess discharge (net production scenario)
        if simulated_consumption < 0:
            simulated_production += abs(simulated_consumption)
            simulated_consumption = 0

    elif strategy == "dynamic_cost_optimization":
    # Dynamic Cost Optimization Strategy

        # Case 1: Charge the battery if the price is below the low threshold
        if total_price_incl_vat_production < config["BATTERY_SIMULATION"].get("DYNAMIC_PRICE_THRESHOLD_LOW", 0.10):
            charge_amount = min(max_charging_rate, charge_maximum - battery_level)
            if charge_amount > 0:
                battery_level += charge_amount * round_trip_efficiency
                energy_loss += charge_amount * (1 - round_trip_efficiency)
                battery_state["total_charged"] += charge_amount
                simulated_production -= charge_amount  # Reduce production used for charging

            else:
                # Case 2: Discharge the battery for consumption if the consumption price is very high
                if hourly_consumption > 0 and total_price_incl_vat_consumption > config["BATTERY_SIMULATION"]["DYNAMIC_PRICE_THRESHOLD_HIGH"]:
                    # Calculate the maximum discharge amount for consumption
                    discharge_amount = min(hourly_consumption, max_discharging_rate, max(0, battery_level - discharge_minimum))
                    if discharge_amount > 0:
                        battery_level -= discharge_amount
                        simulated_consumption -= discharge_amount
                        battery_state["total_discharged"] += discharge_amount

                # Case 3: Discharge the battery for production if the production price is very high
                if total_price_incl_vat_production > config["BATTERY_SIMULATION"]["DYNAMIC_PRICE_THRESHOLD_HIGH"]:
                    # Calculate the maximum discharge amount for production
                    discharge_amount = min(max_discharging_rate, max(0, battery_level - discharge_minimum))
                    if discharge_amount > 0:
                        battery_level -= discharge_amount
                        simulated_production += discharge_amount
                        battery_state["total_discharged"] += discharge_amount

                # Handle excess discharge (net production scenario)
                if simulated_consumption < 0:
                    # Calculate the maximum allowable production based on the battery level
                    max_allowable_production = max(0, battery_level - discharge_minimum)
                    excess_consumption = abs(simulated_consumption)

                    # Only allow production up to the maximum allowable production
                    production_from_battery = min(excess_consumption, max_allowable_production)
                    simulated_production += production_from_battery
                    simulated_consumption += production_from_battery  # Reduce the excess consumption
                    battery_level -= production_from_battery
                    
    # Update the battery state
    battery_state["level"] = battery_level

    # Calculate charge cycles
    usable_capacity = charge_maximum - discharge_minimum
    battery_state["charge_cycles"] = int(battery_state["total_discharged"] // usable_capacity)

    return simulated_consumption, simulated_production, battery_state, energy_loss

def calculate_hourly_energy_prices(base_price, total_annual_consumption, total_annual_production, cumulative_production, salderen):
    """
    Calculate the hourly energy prices for consumption and production.

    Args:
        base_price (float): The base energy price (€/kWh).
        total_annual_consumption (float): Total annual consumption in kWh.
        total_annual_production (float): Total annual production in kWh.
        cumulative_production (float): Cumulative production up to the current hour in kWh.
        salderen (bool): Whether salderen is enabled.

    Returns:
        tuple: (hourly_price_consumption, hourly_price_production)
    """

    # Calculate the hourly energy price for consumption
    hourly_price_consumption = base_price + TAXES["STORAGE_COSTS"] + TAXES["ENERGY_TAX"]
    hourly_price_consumption *= (1 + TAXES["VAT"] / 100)

    # Calculate the hourly energy price for production
    if salderen == False:
        # If salderen is disabled, production price excludes energy tax and VAT for all hours
        hourly_price_production = base_price + TAXES["STORAGE_COSTS_PRODUCTION"]
    else:
        if cumulative_production > total_annual_consumption:
            # For excess production (above total annual consumption), exclude energy tax and VAT
            hourly_price_production = base_price + TAXES["STORAGE_COSTS_PRODUCTION"]
        else:
            # For production within total annual consumption, include energy tax and VAT
            hourly_price_production = base_price + TAXES["STORAGE_COSTS_PRODUCTION"] + TAXES["ENERGY_TAX"]
            hourly_price_production *= (1 + TAXES["VAT"] / 100)
    
    return hourly_price_consumption, hourly_price_production

def calculate_costs(consumption_data, production_data, price_data):
    """Calculate energy costs, income, and total consumption/production, with battery simulation."""
    costs = 0
    income = 0
    total_consumption = 0
    total_production = 0
    total_energy_loss = 0  # Track total energy loss due to battery round-trip efficiency

    # Battery simulation variables
    battery_enabled = config["BATTERY_SIMULATION"]["ENABLE"]
    battery_state = {
        "level": config["BATTERY_SIMULATION"].get("DISCHARGE_LIMIT", 0.1) * config["BATTERY_SIMULATION"]["BATTERY_SIZE_KWH"],
        "total_charged": 0,
        "total_discharged": 0,
        "charge_cycles": 0  # Track the number of charge cycles
    }

    # Get the battery charge strategy
    strategy = config["BATTERY_SIMULATION"].get("BATTERY_CHARGE_STRATEGY", "self-sufficiency")

    # Monthly breakdowns
    monthly_breakdown = {}
    hourly_data = []  # List to store hourly data for the Excel file
    battery_adjusted_costs = 0
    battery_adjusted_income = 0

    # Convert START_DATE and END_DATE to datetime objects
    start_datetime = datetime.strptime(START_DATE, "%Y-%m-%d")
    end_datetime = datetime.strptime(END_DATE, "%Y-%m-%d") + timedelta(days=1)

    # Calculate total annual consumption and production
    total_annual_consumption = sum(consumption_data.values())
    total_annual_production = sum(production_data.values())

    # Initialize cumulative production
    cumulative_production = 0

    for price_entry in price_data:
        # Extract the timestamp and base price
        timestamp_str = price_entry["datum"]
        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%dT%H")

        # Skip entries outside the start and end date range
        if not (start_datetime <= timestamp < end_datetime):
            continue

        # Extract the base price (purchase price excluding VAT)
        base_price = float(price_entry["prijs_excl_belastingen"].replace(",", "."))

        # Get hourly consumption and production values for the timestamp
        hourly_consumption = consumption_data.get(timestamp_str, 0)
        hourly_production = production_data.get(timestamp_str, 0)

        # Update cumulative production
        cumulative_production += hourly_production

        # Calculate hourly energy prices
        hourly_price_consumption, hourly_price_production = calculate_hourly_energy_prices(
            base_price, total_annual_consumption, total_annual_production, cumulative_production, config["PARAMETERS"]["SALDEREN"]
        )

        # Adjust production if STOP_PRODUCTION_NEGATIVE_PRICES is enabled
        adjusted_hourly_production = hourly_production
        if STOP_PRODUCTION_NEGATIVE_PRICES and hourly_price_production < 0:
            debug_print(f"Negative price detected at {timestamp_str}: {hourly_price_production:.2f}. Stopping production.")
            adjusted_hourly_production = 0  # Stop production for this hour

        # Simulate battery behavior if enabled
        if battery_enabled:
            battery_consumption, battery_production, battery_state, energy_loss = simulate_battery(
                hourly_consumption, adjusted_hourly_production, battery_state, config, hourly_price_production, timestamp_str, strategy
            )
            total_energy_loss += energy_loss
            consumption_adjusted = battery_consumption != hourly_consumption
            production_adjusted = battery_production != adjusted_hourly_production
        else:
            battery_consumption = hourly_consumption
            battery_production = adjusted_hourly_production
            consumption_adjusted = False
            production_adjusted = False

        # Accumulate total consumption and production (adjusted values)
        total_consumption += hourly_consumption
        total_production += adjusted_hourly_production

        # Accumulate costs and income (adjusted values)
        costs += hourly_consumption * hourly_price_consumption
        income += adjusted_hourly_production * hourly_price_production

        # Accumulate battery-adjusted costs and income
        battery_adjusted_costs += battery_consumption * hourly_price_consumption
        battery_adjusted_income += battery_production * hourly_price_production

        # Add hourly data for the Excel file
        hourly_data.append({
            "timestamp": timestamp_str,
            "production": hourly_production,
            "adjusted_production": adjusted_hourly_production,
            "consumption": hourly_consumption,
            "simulated_consumption": battery_consumption if battery_enabled else None,
            "simulated_production": battery_production if battery_enabled else None,
            "consumption_adjusted": consumption_adjusted,
            "production_adjusted": production_adjusted,
            "price_consumption": hourly_price_consumption,
            "price_production": hourly_price_production,
            "total_cost_or_income": (
                adjusted_hourly_production * hourly_price_production -
                hourly_consumption * hourly_price_consumption
            ),
            "battery_total_cost_or_income": (
                battery_production * hourly_price_production -
                battery_consumption * hourly_price_consumption
            ) if battery_enabled else None
        })

        # Calculate the month key (e.g., "2024-12")
        month_key = timestamp.strftime("%Y-%m")

        # Initialize monthly breakdown if not already present
        if month_key not in monthly_breakdown:
            monthly_breakdown[month_key] = {
                "costs": 0,
                "income": 0,
                "consumption": 0,
                "production": 0,
                "battery_adjusted_costs": 0,
                "battery_adjusted_income": 0,
                "fixed_supply_costs": FIXED_SUPPLY_COSTS,
                "transport_costs": TRANSPORT_COSTS,
                "energy_tax_compensation": ENERGY_TAX_COMPENSATION
            }

        # Update monthly breakdown
        monthly_breakdown[month_key]["costs"] += hourly_consumption * hourly_price_consumption
        monthly_breakdown[month_key]["income"] += adjusted_hourly_production * hourly_price_production
        monthly_breakdown[month_key]["consumption"] += hourly_consumption
        monthly_breakdown[month_key]["production"] += adjusted_hourly_production
        monthly_breakdown[month_key]["battery_adjusted_costs"] += battery_consumption * hourly_price_consumption
        monthly_breakdown[month_key]["battery_adjusted_income"] += battery_production * hourly_price_production

    # Add fixed monthly costs to the total costs
    for month, data in monthly_breakdown.items():
        data["costs"] += data["fixed_supply_costs"] + data["transport_costs"] + data["energy_tax_compensation"]
        data["battery_adjusted_costs"] += data["fixed_supply_costs"] + data["transport_costs"] + data["energy_tax_compensation"]
        costs += data["fixed_supply_costs"] + data["transport_costs"] + data["energy_tax_compensation"]
        battery_adjusted_costs += data["fixed_supply_costs"] + data["transport_costs"] + data["energy_tax_compensation"]

    return costs, income, total_consumption, total_production, monthly_breakdown, battery_adjusted_costs, battery_adjusted_income, hourly_data, total_energy_loss, battery_state["total_charged"], battery_state["total_discharged"], battery_state["charge_cycles"]

def write_results_to_excel(total_costs, total_income, total_consumption, total_production, monthly_breakdown, battery_adjusted_costs, battery_adjusted_income, hourly_data, total_energy_loss, total_charged, total_discharged, charge_cycles):
    """
    Write the results to an Excel file with multiple sheets: 'settings', 'summary', 'monthly data', and 'hourly data'.
    """
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Add the 'settings' sheet as the first sheet
    settings_sheet = workbook.active
    settings_sheet.title = "settings"

    # Write the settings to the 'settings' sheet
    settings_sheet.append(["Setting", "Value"])
    for key, value in config.items():
        if key == "DATA":
            continue  # Skip the DATA section
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                if isinstance(sub_value, (list, dict)):
                    sub_value = json.dumps(sub_value)  # Convert to JSON string for readability
                settings_sheet.append([f"{key}.{sub_key}", sub_value])
        else:
            if isinstance(value, (list, dict)):
                value = json.dumps(value)  # Convert to JSON string for readability
            settings_sheet.append([key, value])

    # Adjust column widths for the 'settings' sheet
    for column in settings_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        settings_sheet.column_dimensions[column_letter].width = max_length + 2  # Add padding

    # Add the 'summary' sheet explicitly
    summary_sheet = workbook.create_sheet(title="summary")

    # Write the header for the 'summary' sheet
    summary_sheet.append(["Without Battery", "Value", "", "With Battery", "Value"])

    # Calculate total simulated consumption and production for battery data
    total_simulated_consumption = sum(record["simulated_consumption"] for record in hourly_data if record["simulated_consumption"] is not None)
    total_simulated_production = sum(record["simulated_production"] for record in hourly_data if record["simulated_production"] is not None)

    # Calculate the time period in years
    start_date = datetime.strptime(config["PARAMETERS"]["START_DATE"], "%Y-%m-%d")
    end_date = datetime.strptime(config["PARAMETERS"]["END_DATE"], "%Y-%m-%d")
    time_period_years = (end_date - start_date).days / 365.0

    # Calculate the annual savings
    final_cost_non_battery = total_costs - total_income
    final_cost_battery = battery_adjusted_costs - battery_adjusted_income
    annual_savings = (final_cost_non_battery - final_cost_battery) / time_period_years

    # Calculate the payback period for the battery
    battery_price = config["BATTERY_SIMULATION"]["BATTERY_PRICE"]
    payback_period_years = battery_price / annual_savings if annual_savings > 0 else float('inf')

    # Calculate weighted average hourly energy prices
    weighted_avg_price_consumption_non_battery = sum(
        record["price_consumption"] * record["consumption"] for record in hourly_data if record["consumption"] > 0
    ) / total_consumption

    weighted_avg_price_production_non_battery = sum(
        record["price_production"] * record["production"] for record in hourly_data if record["production"] > 0
    ) / total_production

    weighted_avg_price_consumption_battery = sum(
        record["price_consumption"] * record["simulated_consumption"] for record in hourly_data if record["simulated_consumption"] > 0
    ) / total_simulated_consumption

    weighted_avg_price_production_battery = sum(
        record["price_production"] * record["simulated_production"] for record in hourly_data if record["simulated_production"] > 0
    ) / total_simulated_production

    # Define the data for both sections
    non_battery_data = [
        ["Total Costs (€)", total_costs],
        ["Total Income (€)", total_income],
        ["Final Annual Cost (€)", final_cost_non_battery],
        ["Total Consumption (kWh)", total_consumption],
        ["Total Production (kWh)", total_production],
        ["Weighted Avg Price (Consumption €/kWh)", round(weighted_avg_price_consumption_non_battery, 4)],
        ["Weighted Avg Price (Production €/kWh)", round(weighted_avg_price_production_non_battery, 4)]
    ]

    battery_data = [
        ["Total Costs (€)", battery_adjusted_costs],
        ["Total Income (€)", battery_adjusted_income],
        ["Final Annual Cost (€)", final_cost_battery],
        ["Total Consumption (kWh)", total_simulated_consumption],
        ["Total Production (kWh)", total_simulated_production],
        ["Weighted Avg Price (Consumption €/kWh)", round(weighted_avg_price_consumption_battery, 4)],
        ["Weighted Avg Price (Production €/kWh)", round(weighted_avg_price_production_battery, 4)],
        ["Total Energy Loss (kWh)", total_energy_loss],
        ["Total kWh Charged by Battery", total_charged],
        ["Total kWh Discharged by Battery", total_discharged],
        ["Number of Charge Cycles", charge_cycles],
        ["Payback Period (Years)", round(payback_period_years, 2)]
    ]

    # Write the data side by side with an empty column in between
    for i in range(max(len(non_battery_data), len(battery_data))):
        non_battery_row = non_battery_data[i] if i < len(non_battery_data) else ["", ""]
        battery_row = battery_data[i] if i < len(battery_data) else ["", ""]
        summary_sheet.append([non_battery_row[0], non_battery_row[1], "", battery_row[0], battery_row[1]])

    # Adjust column widths for the 'summary' sheet
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        summary_sheet.column_dimensions[column_letter].width = max_length + 2
                                
    # Add the 'monthly data' sheet
    monthly_sheet = workbook.create_sheet(title="monthly data")

    # Write the header for the 'monthly data' sheet
    monthly_sheet.append([
        "Month", 
        "Costs (€)", 
        "Income (€)", 
        "Consumption (kWh)", 
        "Production (kWh)", 
        "Battery-Adjusted Costs (€)", 
        "Battery-Adjusted Income (€)", 
        "Fixed Supply Costs (€)", 
        "Transport Costs (€)", 
        "Energy Tax Compensation (€)", 
        "Net Monthly Costs (€)"
    ])

    # Write the monthly breakdown data
    for month, data in monthly_breakdown.items():
        net_monthly_costs = data["costs"] - data["income"]
        monthly_sheet.append([
            month,
            data["costs"],
            data["income"],
            data["consumption"],
            data["production"],
            data["battery_adjusted_costs"],
            data["battery_adjusted_income"],
            data["fixed_supply_costs"],
            data["transport_costs"],
            data["energy_tax_compensation"],
            net_monthly_costs
        ])

    # Adjust column widths for the 'monthly data' sheet
    for column in monthly_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        monthly_sheet.column_dimensions[column_letter].width = max_length + 2


    # Add the 'hourly data' sheet
    hourly_sheet = workbook.create_sheet(title="hourly data")

    # Write the header for the 'hourly data' sheet
    header = [
        "Date + Hour", 
        "Production (kWh)", 
        "Adjusted Production (kWh)", 
        "Consumption (kWh)", 
        "Hourly Energy Price (Consumption €/kWh)", 
        "Hourly Energy Price (Production €/kWh)", 
        "Total Cost/Income (€)"
    ]
    if config["BATTERY_SIMULATION"]["ENABLE"]:
        header.extend([
            "Simulated Consumption (kWh)", 
            "Simulated Production (kWh)", 
            "Battery-Adjusted Total Cost/Income (€)",
            "Consumption Adjusted", 
            "Production Adjusted"
        ])
    hourly_sheet.append(header)

    # Write the hourly data
    for record in hourly_data:
        row = [
            record["timestamp"],
            record["production"],
            record["adjusted_production"],
            record["consumption"],
            record["price_consumption"],
            record["price_production"],
            record["total_cost_or_income"]
        ]
        if config["BATTERY_SIMULATION"]["ENABLE"]:
            row.extend([
                record["simulated_consumption"],
                record["simulated_production"],
                record["battery_total_cost_or_income"],
                record["consumption_adjusted"],
                record["production_adjusted"]
            ])
        hourly_sheet.append(row)

    # Adjust column widths for the 'hourly data' sheet
    for column in hourly_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        hourly_sheet.column_dimensions[column_letter].width = max_length + 2


    # Add the 'prices' sheet
    prices_sheet = workbook.create_sheet(title="prices")

    # Determine the range for energy prices dynamically
    consumption_prices = [record["price_consumption"] for record in hourly_data]
    production_prices = [record["price_production"] for record in hourly_data]

    # Find the minimum and maximum prices across both consumption and production
    min_price = min(min(consumption_prices), min(production_prices))
    max_price = max(max(consumption_prices), max(production_prices))

    # Create bins for energy prices (20 bins maximum)
    num_bins = 20
    bin_size = (max_price - min_price) / num_bins
    bins = [min_price + i * bin_size for i in range(num_bins + 1)]

    def bin_data(data, bins):
        binned_data = [0] * (len(bins) - 1)
        for value in data:
            for i in range(len(bins) - 1):
                if bins[i] <= value < bins[i + 1]:
                    binned_data[i] += 1
                    break
        return binned_data

    # Bin the data
    binned_consumption = bin_data(consumption_prices, bins)
    binned_production = bin_data(production_prices, bins)

    # Write binned data to the 'prices' sheet
    prices_sheet.append(["Price Range (Consumption)", "kWh Bought", "Price Range (Production)", "kWh Sold"])
    for i in range(num_bins):
        consumption_range = f"{bins[i]:.2f} - {bins[i + 1]:.2f}"
        production_range = f"{bins[i]:.2f} - {bins[i + 1]:.2f}"
        prices_sheet.append([
            consumption_range,
            binned_consumption[i],
            production_range,
            binned_production[i]
        ])

    # Create a bar chart for consumption
    consumption_chart = BarChart()
    consumption_chart.title = "Distribution of Energy Prices (Consumption)"
    consumption_chart.x_axis.title = "Price Range (€/kWh)"
    consumption_chart.y_axis.title = "kWh Bought"
    consumption_data = Reference(prices_sheet, min_col=2, min_row=2, max_row=num_bins + 1)
    consumption_categories = Reference(prices_sheet, min_col=1, min_row=2, max_row=num_bins + 1)
    consumption_chart.add_data(consumption_data, titles_from_data=False)
    consumption_chart.set_categories(consumption_categories)

    # Position the consumption chart in column E
    prices_sheet.add_chart(consumption_chart, "E2")

    # Create a bar chart for production
    production_chart = BarChart()
    production_chart.title = "Distribution of Energy Prices (Production)"
    production_chart.x_axis.title = "Price Range (€/kWh)"
    production_chart.y_axis.title = "kWh Sold"
    production_data = Reference(prices_sheet, min_col=4, min_row=2, max_row=num_bins + 1)
    production_categories = Reference(prices_sheet, min_col=3, min_row=2, max_row=num_bins + 1)
    production_chart.add_data(production_data, titles_from_data=False)
    production_chart.set_categories(production_categories)

    # Position the production chart in column E below the consumption chart
    prices_sheet.add_chart(production_chart, "E20")

    # Generate a human-readable file name
    year = config["PARAMETERS"]["START_DATE"].split("-")[0]  # Extract the year from the start date
    datetime_now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # Current date and time
    salderen = "yes" if config["PARAMETERS"].get("SALDEREN", False) else "no"
    battery_size = config["BATTERY_SIMULATION"]["BATTERY_SIZE_KWH"] if config["BATTERY_SIMULATION"]["ENABLE"] else "no-battery"
    file_name = f"{year}_salderen_{salderen}_battery_{battery_size}kWh_{datetime_now}.xlsx"

    # Save the Excel file
    results_folder = "results"
    os.makedirs(results_folder, exist_ok=True)
    excel_filename = os.path.join(results_folder, file_name)
    workbook.save(excel_filename)

    print(f"Results written to {excel_filename}")

def main():
    # Fetch sensor data from export.json or VictoriaMetrics
    use_export_json = config["DATA"].get("USE_EXPORT_JSON", True)  # Default to using export.json
    sensor_start_date = f"{START_DATE}T00:00:00Z"
    sensor_end_date = f"{END_DATE}T23:59:59Z"

    if use_export_json:
        print("Fetching consumption data from export.json")
        consumption_data = fetch_sensor_data_from_json(config["DATA"].get("EXPORT_JSON_PATH", "data/export.json"), START_DATE, END_DATE, CONSUMPTION_SENSORS)
        print("Consumption data fetched from export.json.")

        print("Fetching production data from export.json")
        production_data = fetch_sensor_data_from_json(config["DATA"].get("EXPORT_JSON_PATH", "data/export.json"), START_DATE, END_DATE, PRODUCTION_SENSORS)
        print("Production data fetched from export.json.")
    else:
        print(f"Fetching consumption data from VictoriaMetrics from {sensor_start_date} to {sensor_end_date}")
        consumption_data = fetch_sensor_data_victoriametrics(
            CONSUMPTION_SENSORS, sensor_start_date, sensor_end_date, "raw_consumption_data.json"
        )
        print("Consumption data fetched and saved to raw_consumption_data.json.")

        print("Fetching production data from VictoriaMetrics")
        production_data = fetch_sensor_data_victoriametrics(
            PRODUCTION_SENSORS, sensor_start_date, sensor_end_date, "raw_production_data.json"
        )
        print("Production data fetched and saved to raw_production_data.json.")

    # Fetch dynamic prices
    price_data = fetch_dynamic_prices(START_DATE, END_DATE)

    # Calculate costs, income, and totals (with and without battery simulation)
    (
        total_costs,
        total_income,
        total_consumption,
        total_production,
        monthly_breakdown,
        battery_adjusted_costs,
        battery_adjusted_income,
        hourly_data,
        total_energy_loss,  # Capture the total energy loss
        total_charged,      # Capture the total kWh charged by the battery
        total_discharged,   # Capture the total kWh discharged by the battery
        charge_cycles       # Capture the number of charge cycles
    ) = calculate_costs(consumption_data, production_data, price_data)

    # Write results to an Excel file
    write_results_to_excel(
        total_costs,
        total_income,
        total_consumption,
        total_production,
        monthly_breakdown,
        battery_adjusted_costs,
        battery_adjusted_income,
        hourly_data,
        total_energy_loss,  # Pass the total energy loss to the Excel writer
        total_charged,      # Pass the total kWh charged to the Excel writer
        total_discharged,   # Pass the total kWh discharged to the Excel writer
        charge_cycles       # Pass the number of charge cycles to the Excel writer
    )

if __name__ == "__main__":
    main()